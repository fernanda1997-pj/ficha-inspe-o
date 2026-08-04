# -*- coding: utf-8 -*-
"""
converter_fichas.py — Geoportal de Inspeção Rodoviária (RTA-MSI / Tocantins)

Lê as fichas de inspeção — tanto de rodovias PAVIMENTADAS quanto NÃO
PAVIMENTADAS (fichas/*.xlsx) — e as linhas de referência de cada S.R.E.
(camadas/R<região>_TRECHOS.shp — as mesmas usadas no geoportal principal em
../web) e gera, para cada região+competência encontrada, um GeoJSON com um
segmento de linha por trecho de KM inspecionado, já cortado na extensão
certa por referenciamento linear (km início/fim ao longo da linha do
S.R.E.). Os dois modelos de ficha têm grupos de condição diferentes:

    pavimentada     -> pavimento · vegetação · elementos de drenagem ·
                        sinalização horizontal · sinalização vertical
    não pavimentada -> condição da plataforma · drenagem superficial

O tipo de cada sheet é detectado automaticamente pelo cabeçalho do primeiro
grupo de condição (não pelo nome do arquivo). Cada grupo guarda os itens
marcados (X) na ficha e um índice de severidade (posição da coluna dentro
do grupo: a 1ª coluna do grupo é sempre a condição boa, as seguintes são
progressivamente piores — é assim que os dois templates são desenhados,
então não precisamos adivinhar o texto exato do rótulo).

Saída (tudo em dados/, consumido pelo index.html sem build):
    dados/insp_<REGIAO>_<AAAA-MM>.js   -- um por região+competência
    dados/manifest.js                  -- lista de todos os anteriores
    relatorio_qualidade.txt            -- SRE não encontrado no shapefile etc.

Rodar:  python converter_fichas.py
Requer: openpyxl, geopandas, shapely
"""
import glob
import json
import os
import re
import sys
import unicodedata
from collections import OrderedDict
from datetime import datetime

import openpyxl
import geopandas as gpd
from shapely.geometry import mapping, LineString
from shapely.ops import substring

BASE = os.path.dirname(os.path.abspath(__file__))
FICHAS_DIR = os.path.join(BASE, 'fichas')
CAMADAS_DIR = os.path.join(BASE, 'camadas')
DADOS_DIR = os.path.join(BASE, 'dados')
EPSG_METRICO = 31982  # SIRGAS 2000 / UTM 22S — mesmo do geoportal principal

MESES_PT = {1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio',
            6: 'Junho', 7: 'Julho', 8: 'Agosto', 9: 'Setembro',
            10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'}

# Grupos de condição, na ordem em que aparecem na ficha (esquerda->direita).
# 'chave_topo' é o texto (normalizado) do cabeçalho mesclado que identifica
# onde o grupo começa/termina na planilha. Cada modelo de ficha (pavimentada
# / não pavimentada) tem o seu próprio conjunto — a chave precisa ser
# específica o bastante pra não casar com o cabeçalho do outro modelo (por
# isso "ELEM. DE DRENAGEM" e não só "DRENAGEM", que também aparece em
# "DRENAGEM SUPERFICIAL" da ficha de não pavimentada).
TEMPLATES = {
    'pavimentada': {
        'nome': 'Pavimentada',
        'chave_deteccao': 'CONDICAO DO PAVIMENTO',
        'grupos': [
            {'id': 'pavimento', 'nome': 'Condição do Pavimento', 'chave_topo': 'CONDICAO DO PAVIMENTO'},
            {'id': 'vegetacao', 'nome': 'Vegetação', 'chave_topo': 'VEGETACAO'},
            {'id': 'drenagem', 'nome': 'Elementos de Drenagem', 'chave_topo': 'ELEM. DE DRENAGEM'},
            {'id': 'sinalizacao_horizontal', 'nome': 'Sinalização Horizontal', 'chave_topo': 'SIN. HOR'},
            {'id': 'sinalizacao_vertical', 'nome': 'Sinalização Vertical', 'chave_topo': 'SIN. VERTICAL'},
        ],
    },
    'nao_pavimentada': {
        'nome': 'Não pavimentada',
        'chave_deteccao': 'CONDICAO DA PLATAFORMA',
        'grupos': [
            {'id': 'plataforma', 'nome': 'Condição da Plataforma', 'chave_topo': 'CONDICAO DA PLATAFORMA'},
            {'id': 'drenagem_superficial', 'nome': 'Drenagem Superficial', 'chave_topo': 'DRENAGEM SUPERFICIAL'},
        ],
    },
}

# Lista combinada (usada só pra escrever o manifest.js com todos os grupos
# possíveis, na ordem em que devem aparecer no painel do mapa).
GRUPOS_TODOS = TEMPLATES['pavimentada']['grupos'] + TEMPLATES['nao_pavimentada']['grupos']

qa_msgs = []


def qa(msg):
    qa_msgs.append(msg)
    print('  [QA]', msg)


def _norm(s):
    """Maiúsculas, sem acento, sem quebra de linha, espaços colapsados."""
    if s is None:
        return ''
    s = str(s)
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r'\s+', ' ', s).strip().upper()
    return s


def _merged_range_de(ws, row, col):
    """Se (row,col) faz parte de uma faixa mesclada, devolve (r1,c1,r2,c2)."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col, rng.max_row, rng.max_col
    return row, col, row, col


def _achar_cabecalho(ws, texto_norm, linhas=range(1, 9)):
    """Acha a célula cujo valor normalizado contém `texto_norm`; devolve (row,col) ou None."""
    for r in linhas:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and texto_norm in _norm(v):
                return r, c
    return None


def parse_ficha(caminho):
    """Lê um arquivo de ficha e devolve lista de dicts (uma por região encontrada)."""
    print(f'Lendo {os.path.basename(caminho)} ...')
    wb = openpyxl.load_workbook(caminho, data_only=True)
    resultados_por_regiao = OrderedDict()

    # aba "Trechos" é só um resumo (LOTE/N./TRECHO/SRE/SUBTRECHOS/EXT/TIPO) —
    # usamos pra ter o nome bonito do trecho e do subtrecho por SRE.
    info_sre = {}
    if 'Trechos' in wb.sheetnames:
        ws = wb['Trechos']
        header_row = None
        for r in range(1, ws.max_row + 1):
            if _norm(ws.cell(row=r, column=1).value) == 'LOTE':
                header_row = r
                break
        if header_row:
            for r in range(header_row + 1, ws.max_row + 1):
                lote = ws.cell(row=r, column=1).value
                n_trecho = ws.cell(row=r, column=2).value
                trecho_nome = ws.cell(row=r, column=3).value
                sre = ws.cell(row=r, column=4).value
                subtrecho = ws.cell(row=r, column=5).value
                ext = ws.cell(row=r, column=6).value
                tipo = ws.cell(row=r, column=7).value
                if not sre:
                    continue
                info_sre[_norm(sre).replace(' ', '')] = {
                    'lote': lote, 'trecho_num': n_trecho, 'trecho_nome': trecho_nome,
                    'subtrecho': subtrecho, 'ext_km': ext, 'tipo': tipo,
                }

    for nome_aba in wb.sheetnames:
        if nome_aba in ('Trechos', 'TT'):
            continue
        ws = wb[nome_aba]

        pos_sre = _achar_cabecalho(ws, 'S.R.E')
        pos_sentido = _achar_cabecalho(ws, 'SENTIDO')
        pos_inicio = _achar_cabecalho(ws, 'INICIO')
        pos_fim = _achar_cabecalho(ws, 'FIM')
        if not (pos_sre and pos_sentido and pos_inicio and pos_fim):
            print(f'  [aviso] aba "{nome_aba}" não parece uma aba de trecho (sem cabeçalho S.R.E/SENTIDO/INICIO/FIM) — pulando')
            continue

        pos_data = _achar_cabecalho(ws, 'DATA:', linhas=range(1, 5))
        pos_regiao = _achar_cabecalho(ws, 'REGIAO:', linhas=range(1, 5))
        pos_ficha = _achar_cabecalho(ws, 'FICHA DE INSPECAO', linhas=range(1, 5))

        competencia_dt = None
        if pos_data:
            r0, c0 = pos_data
            _, _, _, c1 = _merged_range_de(ws, r0, c0)
            for cc in range(c0, min(c0 + 4, ws.max_column + 1)):
                v = ws.cell(row=r0, column=cc).value
                if isinstance(v, datetime):
                    competencia_dt = v
                    break

        regiao_num = None
        if pos_regiao:
            r0, c0 = pos_regiao
            for cc in range(c0, min(c0 + 5, ws.max_column + 1)):
                v = ws.cell(row=r0, column=cc).value
                if isinstance(v, (int, float)):
                    regiao_num = int(v)
                    break

        rodovia_nome = None
        if pos_ficha:
            r0, c0 = pos_ficha
            for cc in range(c0 + 1, min(c0 + 6, ws.max_column + 1)):
                v = ws.cell(row=r0, column=cc).value
                if v:
                    rodovia_nome = str(v).strip()
                    break

        if competencia_dt is None or regiao_num is None:
            print(f'  [aviso] aba "{nome_aba}": não achei DATA e/ou REGIÃO no cabeçalho — pulando')
            continue

        regiao = f'R{regiao_num}'
        competencia = f'{competencia_dt.year:04d}-{competencia_dt.month:02d}'

        # --- qual modelo de ficha é essa aba? (pavimentada x não pavimentada) ---
        r_sre, c_sre = pos_sre
        data_start_row_provisorio = _merged_range_de(ws, r_sre, c_sre)[2] + 1
        tipo_via = None
        for chave_tpl, tpl in TEMPLATES.items():
            if _achar_cabecalho(ws, tpl['chave_deteccao'], linhas=range(1, data_start_row_provisorio)):
                tipo_via = chave_tpl
                break
        if tipo_via is None:
            print(f'  [aviso] aba "{nome_aba}": não reconheci o modelo da ficha (nem pavimentada nem não pavimentada) — pulando')
            continue

        # --- localizar as colunas de cada grupo de condição do modelo detectado ---
        _, c_sentido = pos_sentido
        _, c_inicio = pos_inicio
        _, c_fim = pos_fim
        data_start_row = data_start_row_provisorio  # linha após o cabeçalho de 2 linhas

        grupos_cols = []
        for g in TEMPLATES[tipo_via]['grupos']:
            pos = _achar_cabecalho(ws, g['chave_topo'], linhas=range(1, data_start_row))
            if not pos:
                print(f'  [aviso] aba "{nome_aba}": grupo "{g["nome"]}" não encontrado')
                continue
            r0, c0 = pos
            r1, cc1, r2, cc2 = _merged_range_de(ws, r0, c0)
            cols = list(range(cc1, cc2 + 1))
            labels = [ws.cell(row=r2 + 1, column=cc).value for cc in cols]
            labels = [(re.sub(r'\s+', ' ', str(l)).strip() if l else None) for l in labels]
            grupos_cols.append({**g, 'cols': cols, 'labels': labels})

        # --- ler as linhas de dados ---
        segmentos = []
        sre_atual = None
        sentido_atual = None
        for r in range(data_start_row, ws.max_row + 1):
            v_sre = ws.cell(row=r, column=c_sre).value
            v_sentido = ws.cell(row=r, column=c_sentido).value
            v_ini = ws.cell(row=r, column=c_inicio).value
            v_fim = ws.cell(row=r, column=c_fim).value
            if v_sre not in (None, ''):
                sre_atual = str(v_sre).strip()
            if v_sentido not in (None, ''):
                sentido_atual = str(v_sentido).strip()
            if not isinstance(v_ini, (int, float)) or not isinstance(v_fim, (int, float)):
                continue
            if sre_atual is None:
                continue

            seg = {
                'sre': sre_atual,
                'sentido': sentido_atual,
                'km_ini': round(float(v_ini), 3),
                'km_fim': round(float(v_fim), 3),
                'tipo_via': tipo_via,
            }
            for g in grupos_cols:
                marcados = []
                pior_idx = None
                for i, cc in enumerate(g['cols']):
                    val = ws.cell(row=r, column=cc).value
                    if val not in (None, ''):
                        marcados.append(g['labels'][i] or f'col{i}')
                        pior_idx = i if pior_idx is None else max(pior_idx, i)
                seg[g['id']] = {
                    'marcados': marcados,
                    'severidade': pior_idx,
                    'status': (g['labels'][pior_idx] if pior_idx is not None and g['labels'][pior_idx] else None),
                }
            segmentos.append(seg)

        if not segmentos:
            continue

        # 'trecho' agrupa vários S.R.E. sob o mesmo lote/rodovia — equivale à
        # coluna Id da tabela de atributos do shapefile (R<n>_TRECHOS.shp) e
        # bate com o número da aba da ficha ("202", "206"...). Prioriza o que
        # vem da aba "Trechos" (autoritativo); cai pro nome da aba se faltar.
        def _trecho_de(sre):
            info = info_sre.get(_norm(sre).replace(' ', ''))
            if info and info.get('trecho_num') is not None:
                return info['trecho_num'], info.get('trecho_nome') or rodovia_nome
            return nome_aba, rodovia_nome

        segs_com_trecho = []
        for s in segmentos:
            trecho_num, trecho_nome = _trecho_de(s['sre'])
            segs_com_trecho.append({**s, 'rodovia': rodovia_nome, 'aba': nome_aba,
                                     'trecho_num': trecho_num, 'trecho_nome': trecho_nome})
        resultados_por_regiao.setdefault(regiao, {}).setdefault(competencia, []).extend(segs_com_trecho)

    saida = []
    for regiao, por_comp in resultados_por_regiao.items():
        for competencia, segs in por_comp.items():
            saida.append({'regiao': regiao, 'competencia': competencia, 'segmentos': segs})
    return saida


def carregar_linhas_regiao(regiao):
    """Carrega camadas/<regiao>_TRECHOS.shp; devolve dict SRE(normalizado) -> dict com geometria."""
    caminho = os.path.join(CAMADAS_DIR, f'{regiao}_TRECHOS.shp')
    if not os.path.exists(caminho):
        qa(f'{regiao}: shapefile {os.path.basename(caminho)} não encontrado em camadas/ — segmentos dessa região foram ignorados')
        return {}
    gdf_m = gpd.read_file(caminho)
    if gdf_m.crs is None:
        gdf_m = gdf_m.set_crs(EPSG_METRICO)
    else:
        gdf_m = gdf_m.to_crs(EPSG_METRICO)
    gdf_deg = gdf_m.to_crs(4326)

    col_sre = next((c for c in gdf_m.columns if _norm(c) in ('SRE',)), None)
    if col_sre is None:
        qa(f'{regiao}: shapefile sem coluna SRE reconhecível (colunas: {list(gdf_m.columns)})')
        return {}

    linhas = {}
    for i in range(len(gdf_m)):
        sre_raw = gdf_m.iloc[i][col_sre]
        if sre_raw is None:
            continue
        sre_key = _norm(sre_raw).replace(' ', '')
        if not sre_key:
            continue
        geom_m = gdf_m.iloc[i].geometry
        if geom_m is None or geom_m.is_empty:
            continue
        if geom_m.geom_type == 'MultiLineString':
            # As partes não se tocam exatamente (emenda de digitalização) —
            # concatena na ordem em que estão no shapefile (é assim que elas
            # seguem o traçado da rodovia), invertendo cada parte se for
            # preciso pra manter a continuidade com a parte anterior. Isso
            # preserva o comprimento total (bate com EXT_REAL) em vez de
            # descartar pedaços, mas se o vão entre duas partes for grande,
            # o traço vira uma linha reta nesse trecho (por isso avisamos).
            partes = list(geom_m.geoms)
            maior_vao = 0.0
            coords = list(partes[0].coords)
            for p in partes[1:]:
                pc = list(p.coords)
                ultimo = coords[-1]
                d_ini = (ultimo[0] - pc[0][0]) ** 2 + (ultimo[1] - pc[0][1]) ** 2
                d_fim = (ultimo[0] - pc[-1][0]) ** 2 + (ultimo[1] - pc[-1][1]) ** 2
                if d_fim < d_ini:
                    pc = pc[::-1]
                vao = min(d_ini, d_fim) ** 0.5
                maior_vao = max(maior_vao, vao)
                coords.extend(pc)
            geom_m = LineString(coords)
            qa(f'{regiao}/{sre_raw}: geometria em {len(partes)} partes desconexas no shapefile — '
               f'concatenadas na ordem original (maior vão entre partes: {round(maior_vao)} m'
               f'{" — CONFERIR o shapefile, pode estar faltando um pedaço do traçado" if maior_vao > 200 else ""})')
        if sre_key in linhas:
            qa(f'{regiao}: SRE {sre_raw!r} duplicado no shapefile — usando a primeira ocorrência')
            continue

        # Descobre se o vértice inicial da linha corresponde ao km 0 (INÍCIO)
        # ou ao km final (FIM) do trecho, comparando com os atributos
        # X_LONG/Y_LAT (início) e X_FIM_LONG/Y_FIM_LAT (fim) — ambos em graus.
        # Se estiver invertida, vira a linha antes de cortar, senão
        # km_ini/km_fim ficam trocados. Reprojeta só o 1º vértice (já
        # fundido/reduzido a LineString única) pra graus, na hora.
        ini_lonlat = (gdf_m.iloc[i].get('X_LONG'), gdf_m.iloc[i].get('Y_LAT'))
        fim_lonlat = (gdf_m.iloc[i].get('X_FIM_LONG'), gdf_m.iloc[i].get('Y_FIM_LAT'))
        geom_deg = gpd.GeoSeries([geom_m], crs=EPSG_METRICO).to_crs(4326).iloc[0]
        p0 = geom_deg.coords[0]
        d0 = _dist2(p0, ini_lonlat)
        d1 = _dist2(p0, fim_lonlat)
        if d0 is not None and d1 is not None and d1 < d0:
            geom_m = LineString(list(geom_m.coords)[::-1])
        elif d0 is None:
            qa(f'{regiao}/{sre_raw}: sem X_LONG/Y_LAT no shapefile pra checar o sentido da linha — assumindo a ordem original')

        linhas[sre_key] = {
            'geom_m': geom_m,
            'comprimento_m': geom_m.length,
            'ext_real_km': gdf_m.iloc[i].get('EXT_REAL'),
        }
    return linhas


def _dist2(a, b):
    if a[0] is None or b[0] is None or a[1] is None or b[1] is None:
        return None
    return (a[0] - b[0]) ** 2 + (a[1] - b[1]) ** 2


def cortar_segmento(linha_info, km_ini, km_fim, regiao, sre):
    geom = linha_info['geom_m']
    comprimento = linha_info['comprimento_m']

    d_ini_m = max(0.0, min(km_ini * 1000.0, comprimento))
    d_fim_m = max(0.0, min(km_fim * 1000.0, comprimento))
    if d_fim_m <= d_ini_m:
        d_fim_m = min(comprimento, d_ini_m + 0.5)

    cortado = substring(geom, d_ini_m, d_fim_m)
    if cortado.is_empty or cortado.geom_type != 'LineString' or len(cortado.coords) < 2:
        return None
    return cortado


def js_string(valor):
    return json.dumps(valor, ensure_ascii=False)


def gerar_regioes():
    """Lê camadas/R<n>_REGIÃO.shp (contorno de cada região) e escreve dados/regioes.js —
    usado só pra desenhar o limite da região selecionada no mapa (contexto visual)."""
    caminhos = sorted(glob.glob(os.path.join(CAMADAS_DIR, 'R*_REGIÃO.shp')))
    if not caminhos:
        qa('Nenhum shapefile R*_REGIÃO.shp encontrado em camadas/ — sem contorno de região no mapa')
        return
    features = []
    for caminho in caminhos:
        m = re.match(r'R(\d+)_', os.path.basename(caminho))
        if not m:
            continue
        regiao = f'R{int(m.group(1))}'
        gdf = gpd.read_file(caminho)
        gdf = gdf.to_crs(4326) if gdf.crs else gdf.set_crs(4326)
        for geom in gdf.geometry:
            if geom is None or geom.is_empty:
                continue
            features.append({'type': 'Feature', 'properties': {'regiao': regiao},
                              'geometry': mapping(geom)})
    geojson = {'type': 'FeatureCollection', 'features': features}
    with open(os.path.join(DADOS_DIR, 'regioes.js'), 'w', encoding='utf-8') as f:
        f.write('// Gerado por converter_fichas.py — não editar à mão\n')
        f.write(f'window.DADOS_REGIOES = {json.dumps(geojson, ensure_ascii=False)};\n')
    print(f'dados/regioes.js escrito com {len(features)} região(ões).')


def main():
    os.makedirs(DADOS_DIR, exist_ok=True)
    gerar_regioes()
    arquivos = sorted(glob.glob(os.path.join(FICHAS_DIR, '*.xlsx')))
    if not arquivos:
        print(f'Nenhuma ficha .xlsx encontrada em {FICHAS_DIR}')
        return

    por_regiao_competencia = OrderedDict()  # (regiao,competencia) -> lista de segmentos crus

    for caminho in arquivos:
        try:
            saida = parse_ficha(caminho)
        except Exception as e:
            qa(f'{os.path.basename(caminho)}: falha ao ler ({e}) — arquivo ignorado')
            continue
        for bloco in saida:
            chave = (bloco['regiao'], bloco['competencia'])
            por_regiao_competencia.setdefault(chave, []).extend(bloco['segmentos'])

    cache_linhas = {}
    manifest = []

    for (regiao, competencia), segmentos in por_regiao_competencia.items():
        print(f'\n=== {regiao} — competência {competencia} — {len(segmentos)} segmento(s) na ficha ===')
        if regiao not in cache_linhas:
            cache_linhas[regiao] = carregar_linhas_regiao(regiao)
        linhas = cache_linhas[regiao]

        features = []
        sre_nao_achados = set()
        ultimo_fim_por_sre = {}

        for seg in segmentos:
            sre_key = _norm(seg['sre']).replace(' ', '')
            linha_info = linhas.get(sre_key)
            if linha_info is None:
                sre_nao_achados.add(seg['sre'])
                continue
            geom_cortada = cortar_segmento(linha_info, seg['km_ini'], seg['km_fim'], regiao, seg['sre'])
            if geom_cortada is None:
                qa(f'{regiao}/{seg["sre"]}: km {seg["km_ini"]}-{seg["km_fim"]} virou geometria vazia — pulado')
                continue
            ultimo_fim_por_sre[sre_key] = max(ultimo_fim_por_sre.get(sre_key, 0), seg['km_fim'])

            props = {
                'regiao': regiao,
                'competencia': competencia,
                'competencia_label': f"{MESES_PT[int(competencia.split('-')[1])]}/{competencia.split('-')[0]}",
                'sre': seg['sre'],
                'sentido': seg['sentido'],
                'tipo_via': seg.get('tipo_via'),
                'rodovia': seg.get('rodovia'),
                'trecho_num': seg.get('trecho_num'),
                'trecho_nome': seg.get('trecho_nome'),
                'km_ini': seg['km_ini'],
                'km_fim': seg['km_fim'],
            }
            # só grava o grupo se ele existir na ficha dessa via (pavimentada
            # tem 5 grupos, não pavimentada tem 2 — o mapa filtra cada camada
            # pelas features que realmente têm aquela chave em properties).
            for g in GRUPOS_TODOS:
                info_g = seg.get(g['id'])
                if info_g is None:
                    continue
                props[g['id']] = {
                    'status': info_g.get('status'),
                    'severidade': info_g.get('severidade'),
                    'marcados': info_g.get('marcados', []),
                }

            features.append({
                'type': 'Feature',
                'geometry': {'type': 'LineString', 'coords_m': list(geom_cortada.coords)},
                'properties': props,
            })

        if sre_nao_achados:
            qa(f'{regiao}/{competencia}: {len(sre_nao_achados)} S.R.E. da ficha não encontrado(s) no shapefile: {sorted(sre_nao_achados)}')

        for sre_key, fim in ultimo_fim_por_sre.items():
            ext_real = linhas.get(sre_key, {}).get('ext_real_km')
            if ext_real and abs(fim - ext_real) > max(0.3, ext_real * 0.03):
                qa(f'{regiao}/{competencia}: SRE {sre_key} — km final inspecionado ({fim}) difere da extensão do shapefile ({round(ext_real,3)})')

        if not features:
            print(f'  nenhum segmento georreferenciado para {regiao}/{competencia} — nada gerado')
            continue

        # reprojeta em lote (métrico -> WGS84) pra ficar rápido
        transformer = None
        try:
            from pyproj import Transformer
            transformer = Transformer.from_crs(EPSG_METRICO, 4326, always_xy=True)
        except Exception as e:
            qa(f'pyproj indisponível ({e}) — abortando')
            return

        for feat in features:
            coords_m = feat['geometry'].pop('coords_m')
            xs = [p[0] for p in coords_m]
            ys = [p[1] for p in coords_m]
            lons, lats = transformer.transform(xs, ys)
            feat['geometry']['coordinates'] = [[round(lo, 6), round(la, 6)] for lo, la in zip(lons, lats)]

        geojson = {'type': 'FeatureCollection', 'features': features}
        nome_arquivo = f'insp_{regiao}_{competencia}.js'
        var_nome = f"DADOS_INSPECAO_{regiao}_{competencia.replace('-', '_')}"
        conteudo = (
            f"// Gerado por converter_fichas.py — não editar à mão\n"
            f"window.DADOS_INSPECAO = window.DADOS_INSPECAO || {{}};\n"
            f"window.DADOS_INSPECAO[{js_string(regiao)}] = window.DADOS_INSPECAO[{js_string(regiao)}] || {{}};\n"
            f"window.DADOS_INSPECAO[{js_string(regiao)}][{js_string(competencia)}] = {json.dumps(geojson, ensure_ascii=False)};\n"
        )
        with open(os.path.join(DADOS_DIR, nome_arquivo), 'w', encoding='utf-8') as f:
            f.write(conteudo)
        print(f'  -> dados/{nome_arquivo} ({len(features)} segmentos)')

        ext_total = sum(
            (f['properties']['km_fim'] - f['properties']['km_ini']) for f in features
        )
        manifest.append({
            'regiao': regiao,
            'competencia': competencia,
            'competencia_label': features[0]['properties']['competencia_label'],
            'arquivo': nome_arquivo,
            'n_segmentos': len(features),
            'ext_km': round(ext_total, 1),
        })

    manifest.sort(key=lambda m: (m['regiao'], m['competencia']))
    grupos_saida = [{'id': g['id'], 'nome': g['nome']} for g in GRUPOS_TODOS]
    with open(os.path.join(DADOS_DIR, 'manifest.js'), 'w', encoding='utf-8') as f:
        f.write('// Gerado por converter_fichas.py — não editar à mão\n')
        f.write(f'window.MANIFEST_INSPECAO = {json.dumps(manifest, ensure_ascii=False, indent=2)};\n')
        f.write(f'window.GRUPOS_INSPECAO = {json.dumps(grupos_saida, ensure_ascii=False, indent=2)};\n')
    print(f'\ndados/manifest.js escrito com {len(manifest)} conjunto(s) região/competência.')

    with open(os.path.join(BASE, 'relatorio_qualidade.txt'), 'w', encoding='utf-8') as f:
        if qa_msgs:
            f.write('\n'.join(qa_msgs) + '\n')
        else:
            f.write('Nenhum problema encontrado.\n')
    print(f'relatorio_qualidade.txt: {len(qa_msgs)} observação(ões).')


if __name__ == '__main__':
    main()
