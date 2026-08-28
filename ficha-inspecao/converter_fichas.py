# -*- coding: utf-8 -*-
"""
converter_fichas.py — Geoportal de Inspeção Rodoviária (RTA-MSI / Tocantins)

Lê as fichas de inspeção — tanto de rodovias PAVIMENTADAS quanto NÃO
PAVIMENTADAS (fichas/*.xlsx) — e as linhas de referência de cada S.R.E.
(camadas/R<região>_TRECHOS.shp — as mesmas usadas no geoportal principal em
../web) e gera, para cada região+competência encontrada, um GeoJSON com um
segmento de linha por trecho de KM inspecionado, já cortado na extensão
certa por referenciamento linear (km início/fim ao longo da linha do
S.R.E.). Os dois modelos de ficha têm grupos de condição diferentes:

    pavimentada     -> pavimento · vegetação · elementos de drenagem ·
                        sinalização horizontal · sinalização vertical
    não pavimentada -> condição da plataforma · drenagem superficial

O tipo de cada sheet é detectado automaticamente pelo cabeçalho do primeiro
grupo de condição (não pelo nome do arquivo). Cada grupo guarda os itens
marcados (X) na ficha e um índice de severidade (posição da coluna dentro
do grupo: a 1ª coluna do grupo é sempre a condição boa, as seguintes são
progressivamente piores — é assim que os dois templates são desenhados,
então não precisamos adivinhar o texto exato do rótulo).

Saída (tudo em dados/, consumido pelo index.html sem build):
    dados/insp_<REGIAO>_<AAAA-MM>.js   -- um por região+competência
    dados/manifest.js                  -- lista de todos os anteriores
    relatorio_qualidade.txt            -- SRE não encontrado no shapefile etc.

Rodar:  python converter_fichas.py
Requer: openpyxl, geopandas, shapely
"""
import glob
import json
import math
import os
import re
import sys
import unicodedata
from collections import OrderedDict
from datetime import datetime

import openpyxl
import geopandas as gpd
from shapely.geometry import mapping, LineString
from shapely.ops import substring

BASE = os.path.dirname(os.path.abspath(__file__))
FICHAS_DIR = os.path.join(BASE, 'fichas')
CAMADAS_DIR = os.path.join(BASE, '..', 'camadas')  # compartilhado com ordens-servico/
DADOS_DIR = os.path.join(BASE, 'dados')
EPSG_METRICO = 31982  # SIRGAS 2000 / UTM 22S — mesmo do geoportal principal

# Planilha de controle dos Pontos Críticos vive no OUTRO projeto (o geoportal
# principal Folium, pasta "web"), não em "web - fichas" — é lá que a usuária
# edita/atualiza o status mês a mês. Lida direto de lá (path absoluto) pra não
# correr o risco de rodar com uma cópia desatualizada; só as geometrias
# (camadas/R<n>_Pontos_Criticos.shp) é que foram copiadas pra web - fichas/camadas
# porque essas praticamente não mudam.
WEB_PRINCIPAL_DIR = os.path.join(BASE, '..', '..', 'web')
PLANILHA_PONTOS_CRITICOS = os.path.join(WEB_PRINCIPAL_DIR, 'pontos criticos', 'Controle Pontos Críticos .xlsx')

MESES_PT = {1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio',
            6: 'Junho', 7: 'Julho', 8: 'Agosto', 9: 'Setembro',
            10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'}
# nome (maiúsculo, sem acento) -> número — usado no fallback de competência
# a partir do nome do arquivo, quando a ficha não tem a célula DATA: preenchida.
NUM_DO_MES_PT = {unicodedata.normalize('NFKD', v).encode('ascii', 'ignore').decode('ascii').upper(): k
                 for k, v in MESES_PT.items()}

# Grupos de condição, na ordem em que aparecem na ficha (esquerda->direita).
# 'chave_topo' é o texto (normalizado) do cabeçalho mesclado que identifica
# onde o grupo começa/termina na planilha. Cada modelo de ficha (pavimentada
# / não pavimentada) tem o seu próprio conjunto — a chave precisa ser
# específica o bastante pra não casar com o cabeçalho do outro modelo (por
# isso "ELEM. DE DRENAGEM" e não só "DRENAGEM", que também aparece em
# "DRENAGEM SUPERFICIAL" da ficha de não pavimentada).
TEMPLATES = {
    'pavimentada': {
        'nome': 'Pavimentada',
        'chave_deteccao': 'CONDICAO DO PAVIMENTO',
        'grupos': [
            {'id': 'pavimento', 'nome': 'Condição do Pavimento', 'chave_topo': 'CONDICAO DO PAVIMENTO'},
            {'id': 'vegetacao', 'nome': 'Vegetação', 'chave_topo': 'VEGETACAO'},
            {'id': 'drenagem', 'nome': 'Elementos de Drenagem', 'chave_topo': 'ELEM. DE DRENAGEM'},
            {'id': 'sinalizacao_horizontal', 'nome': 'Sinalização Horizontal', 'chave_topo': 'SIN. HOR'},
            {'id': 'sinalizacao_vertical', 'nome': 'Sinalização Vertical', 'chave_topo': 'SIN. VERTICAL'},
        ],
    },
    'nao_pavimentada': {
        'nome': 'Não pavimentada',
        'chave_deteccao': 'CONDICAO DA PLATAFORMA',
        'grupos': [
            {'id': 'plataforma', 'nome': 'Condição da Plataforma', 'chave_topo': 'CONDICAO DA PLATAFORMA'},
            {'id': 'drenagem_superficial', 'nome': 'Drenagem Superficial', 'chave_topo': 'DRENAGEM SUPERFICIAL'},
        ],
    },
}

# Lista combinada (usada só pra escrever o manifest.js com todos os grupos
# possíveis, na ordem em que devem aparecer no painel do mapa).
GRUPOS_TODOS = TEMPLATES['pavimentada']['grupos'] + TEMPLATES['nao_pavimentada']['grupos']

# Maior severidade possível de cada grupo (posição da última coluna) — usado
# só pra normalizar 0–1 na hora de calcular o I.C.M./I.C.M.N.P. geral.
MAX_SEVERIDADE = {
    'pavimento': 4, 'vegetacao': 1, 'drenagem': 2,
    'sinalizacao_horizontal': 2, 'sinalizacao_vertical': 2,
    'plataforma': 3, 'drenagem_superficial': 2,
}

# Faixas do índice geral (média das severidades normalizadas dos grupos
# presentes no segmento). Combinado com a usuária: 25% em 25%.
FAIXAS_ICM = [(0.25, 'Bom'), (0.50, 'Regular'), (0.75, 'Ruim'), (1.01, 'Péssimo')]


def calcular_icm(props):
    """Índice geral (I.C.M./I.C.M.N.P.): média das severidades normalizadas
    (0–1) dos grupos presentes no segmento, mapeada em Bom/Regular/Ruim/
    Péssimo (25% em 25%). Sem nenhum grupo marcado -> 'Sem Informação'."""
    valores = []
    for grupo_id, maximo in MAX_SEVERIDADE.items():
        info = props.get(grupo_id)
        if info and info.get('severidade') is not None:
            valores.append(info['severidade'] / maximo)
    if not valores:
        return {'classe': 'Sem Informação', 'valor': None}
    media = sum(valores) / len(valores)
    for limite, nome in FAIXAS_ICM:
        if media <= limite:
            return {'classe': nome, 'valor': round(media, 4)}
    return {'classe': 'Péssimo', 'valor': round(media, 4)}


qa_msgs = []


def qa(msg):
    qa_msgs.append(msg)
    print('  [QA]', msg)


def _norm(s):
    """Maiúsculas, sem acento, sem quebra de linha, espaços colapsados."""
    if s is None:
        return ''
    s = str(s)
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r'\s+', ' ', s).strip().upper()
    return s


def _merged_range_de(ws, row, col):
    """Se (row,col) faz parte de uma faixa mesclada, devolve (r1,c1,r2,c2)."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col, rng.max_row, rng.max_col
    return row, col, row, col


def _achar_cabecalho(ws, texto_norm, linhas=range(1, 9), exato=False):
    """Acha a célula cujo valor normalizado contém `texto_norm`; devolve (row,col) ou None.
    `exato=True` exige igualdade (não só conter) — usado pro rótulo "S.R.E."
    da coluna, que senão casa de propósito errado com a frase "...· 5 S.R.E."
    da linha de descrição do trecho (modelo novo, jul/2026)."""
    for r in linhas:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not v:
                continue
            v_norm = _norm(v)
            casou = (v_norm == texto_norm) if exato else (texto_norm in v_norm)
            if casou:
                return r, c
    return None


def reconstruir_run_quebrado(sre_key, linhas_run, trecho_ordem, info_sre):
    """Tenta corrigir um 'run' de linhas com o mesmo S.R.E. escrito cujo KM
    ultrapassa muito a extensão oficial dele — sinal de que o código não foi
    atualizado quando a estrada passou pro próximo S.R.E. oficial (comum em
    fichas preenchidas em campo sem muita disciplina).

    Ignora os números absolutos de KM escritos (alguns resetam a cada S.R.E.,
    outros continuam — sem padrão): funde linhas consecutivas com KM
    idêntico (mesma posição, 2 faixas/sentidos — fica com a PIOR marca de
    cada grupo) e distribui a largura medida dentro do "orçamento" oficial
    (esse S.R.E. + quantos seguintes do mesmo trecho forem precisos).

    Só devolve resultado se sair uma leitura "limpa" — sem KM pra trás e sem
    precisar espremer/esticar mais que ~80%. Senão devolve None (quem chamou
    decide deixar como está + avisar que precisa revisão manual)."""
    info = info_sre.get(sre_key)
    if not info or info.get('trecho_num') is None:
        return None
    trecho_key = str(info['trecho_num']).strip()
    lista_oficial = trecho_ordem.get(trecho_key)
    if not lista_oficial:
        return None
    pos_inicio = next((i for i, (s, _) in enumerate(lista_oficial) if s == sre_key), None)
    if pos_inicio is None:
        return None

    for l in linhas_run:
        if l['km_fim'] < l['km_ini'] - 1e-6:
            return None  # km pra trás dentro do run — não confiável

    campos_fixos = ('sre', 'sentido', 'km_ini', 'km_fim', 'tipo_via')
    grupos_ids = [k for k in linhas_run[0].keys() if k not in campos_fixos]

    passos = []
    i = 0
    while i < len(linhas_run):
        j = i + 1
        while (j < len(linhas_run) and linhas_run[j]['km_ini'] == linhas_run[i]['km_ini']
               and linhas_run[j]['km_fim'] == linhas_run[i]['km_fim']):
            j += 1
        grupo_dup = linhas_run[i:j]
        largura = linhas_run[i]['km_fim'] - linhas_run[i]['km_ini']
        if largura <= 0:
            largura = 0.001
        marcas_final = {}
        for g in grupos_ids:
            candidatos = [x[g] for x in grupo_dup if x.get(g) is not None]
            marcas_final[g] = (max(candidatos, key=lambda ig: (ig.get('severidade') if ig.get('severidade') is not None else -1))
                                if candidatos else None)
        passos.append({'largura': largura, 'marcas': marcas_final})
        i = j

    largura_total = sum(p['largura'] for p in passos)
    if largura_total <= 0:
        return None

    orcamento = []
    soma = 0.0
    for s, ext in lista_oficial[pos_inicio:]:
        if len(orcamento) >= 5:
            break  # sinal de que o run não tem nada a ver com esse trecho — não confiável
        orcamento.append((s, ext))
        soma += ext
        if soma >= largura_total - 1e-9:
            break
    if soma <= 0:
        return None
    if soma >= largura_total - 1e-9:
        # orçamento oficial dá conta (com folga ou justo) — não precisa
        # comprimir nada, só sobra um pedacinho do último S.R.E. sem
        # preencher (significa que aquele trecho não foi 100% inspecionado,
        # o que é normal).
        fator = 1.0
    else:
        # usou TODO o resto oficial do trecho e ainda faltou — só nesse
        # caso precisa esticar (e só aceita até 80% de esticamento).
        fator = largura_total / soma
    if fator > 1.8:
        return None  # precisaria esticar demais pra caber — não confiável

    EPS = 1e-6
    saida = []
    idx_orc = 0
    usado_no_sre = 0.0
    sentido = linhas_run[0].get('sentido')
    tipo_via = linhas_run[0].get('tipo_via')
    for passo in passos:
        restante = passo['largura'] * fator
        while restante > EPS:
            no_ultimo = idx_orc >= len(orcamento) - 1
            sre_atual, ext_atual = orcamento[min(idx_orc, len(orcamento) - 1)]
            espaco_livre = ext_atual - usado_no_sre
            usar = restante if no_ultimo else min(restante, max(espaco_livre, 0.0))
            if usar <= EPS and not no_ultimo:
                idx_orc += 1
                usado_no_sre = 0.0
                continue
            novo = {'sre': sre_atual, 'sentido': sentido, 'tipo_via': tipo_via,
                    'km_ini': round(usado_no_sre, 3), 'km_fim': round(usado_no_sre + usar, 3)}
            for g in grupos_ids:
                novo[g] = passo['marcas'][g]
            saida.append(novo)
            usado_no_sre += usar
            restante -= usar
            if (not no_ultimo) and usado_no_sre >= ext_atual - EPS:
                idx_orc += 1
                usado_no_sre = 0.0
    return saida


def corrigir_sre_travado(segmentos, info_sre, trecho_ordem, regiao, nome_aba):
    """Agrupa `segmentos` (de UMA aba) em 'runs' de S.R.E. repetido; pros
    runs cujo KM passa muito da extensão oficial, tenta reconstruir (ver
    `reconstruir_run_quebrado`) — só substitui se sair limpo, senão deixa
    como estava e avisa que precisa revisão manual."""
    if not segmentos:
        return segmentos
    runs = []
    for s in segmentos:
        if runs and runs[-1][0] == s['sre']:
            runs[-1][1].append(s)
        else:
            runs.append((s['sre'], [s]))

    saida = []
    for sre_original, linhas_run in runs:
        sre_key = _norm(sre_original).replace(' ', '')
        info = info_sre.get(sre_key)
        oficial = info.get('ext_km') if info else None
        span = max(l['km_fim'] for l in linhas_run) - min(l['km_ini'] for l in linhas_run)
        if oficial and isinstance(oficial, (int, float)) and span > oficial * 1.5 + 0.5:
            corrigido = reconstruir_run_quebrado(sre_key, linhas_run, trecho_ordem, info_sre)
            if corrigido is not None:
                qa(f'{regiao}/{nome_aba}: SRE {sre_original!r} tinha KM até {span:.1f} km (oficial {oficial:.2f} km, '
                   f'código não atualizado a tempo) — RECONSTRUÍDO automaticamente em {len(corrigido)} sub-trecho(s) '
                   f'usando a ordem/extensão oficial da aba "Trechos"')
                saida.extend(corrigido)
                continue
            else:
                qa(f'{regiao}/{nome_aba}: SRE {sre_original!r} tinha KM até {span:.1f} km (oficial só {oficial:.2f} km) — '
                   f'PENDENTE DE REVISÃO MANUAL (KM fora de ordem ou incompatível demais com o orçamento oficial do '
                   f'trecho pra eu reconstruir com confiança) — mantido como estava na ficha')
        saida.extend(linhas_run)
    return saida


def parse_ficha(caminho):
    """Lê um arquivo de ficha e devolve lista de dicts (uma por região encontrada)."""
    print(f'Lendo {os.path.basename(caminho)} ...')
    wb = openpyxl.load_workbook(caminho, data_only=True)
    resultados_por_regiao = OrderedDict()

    # aba "Trechos" é só um resumo (LOTE/N./TRECHO/SRE/SUBTRECHOS/EXT/TIPO) —
    # usamos pra ter o nome bonito do trecho e do subtrecho por SRE, e também
    # a ORDEM oficial + extensão de cada S.R.E. dentro do trecho (usada pra
    # reconstruir abas onde o código do S.R.E. não foi atualizado a tempo —
    # ver `reconstruir_run_quebrado`).
    info_sre = {}
    trecho_ordem = OrderedDict()  # trecho_num(str) -> [(sre_key, ext_km), ...]
    if 'Trechos' in wb.sheetnames:
        ws = wb['Trechos']
        header_row = None
        for r in range(1, ws.max_row + 1):
            if _norm(ws.cell(row=r, column=1).value) == 'LOTE':
                header_row = r
                break
        if header_row:
            for r in range(header_row + 1, ws.max_row + 1):
                lote = ws.cell(row=r, column=1).value
                n_trecho = ws.cell(row=r, column=2).value
                trecho_nome = ws.cell(row=r, column=3).value
                sre = ws.cell(row=r, column=4).value
                subtrecho = ws.cell(row=r, column=5).value
                ext = ws.cell(row=r, column=6).value
                tipo = ws.cell(row=r, column=7).value
                if not sre:
                    continue
                sre_key = _norm(sre).replace(' ', '')
                info_sre[sre_key] = {
                    'lote': lote, 'trecho_num': n_trecho, 'trecho_nome': trecho_nome,
                    'subtrecho': subtrecho, 'ext_km': ext, 'tipo': tipo,
                }
                if n_trecho is not None and isinstance(ext, (int, float)):
                    trecho_ordem.setdefault(str(n_trecho).strip(), []).append((sre_key, float(ext)))

    for nome_aba in wb.sheetnames:
        if nome_aba in ('Trechos', 'TT'):
            continue
        ws = wb[nome_aba]

        pos_sre = _achar_cabecalho(ws, 'S.R.E.', exato=True)
        # Restringe SENTIDO/INÍCIO/FIM à MESMA linha do "S.R.E." (quando achado)
        # — senão "FIM" casa de propósito errado com o texto de instrução
        # ("...até o FIM de CADA S.R.E...") que vem ANTES do cabeçalho de
        # verdade no modelo novo (jul/2026). Sem pos_sre, cai pro range
        # amplo de antes (aba provavelmente não é de trecho mesmo).
        linhas_cab = [pos_sre[0]] if pos_sre else range(1, 9)
        # 'SENTIDO' no modelo antigo, 'SENT.' no modelo novo — aceita os dois.
        pos_sentido = (_achar_cabecalho(ws, 'SENTIDO', linhas=linhas_cab)
                       or _achar_cabecalho(ws, 'SENT.', linhas=linhas_cab))
        pos_inicio = _achar_cabecalho(ws, 'INICIO', linhas=linhas_cab)
        pos_fim = _achar_cabecalho(ws, 'FIM', linhas=linhas_cab)
        if not (pos_sre and pos_sentido and pos_inicio and pos_fim):
            print(f'  [aviso] aba "{nome_aba}" não parece uma aba de trecho (sem cabeçalho S.R.E/SENTIDO/INICIO/FIM) — pulando')
            continue

        pos_data = _achar_cabecalho(ws, 'DATA:', linhas=range(1, 5))
        pos_regiao = _achar_cabecalho(ws, 'REGIAO:', linhas=range(1, 5))
        pos_ficha = _achar_cabecalho(ws, 'FICHA DE INSPECAO', linhas=range(1, 5))

        competencia_dt = None
        if pos_data:
            r0, c0 = pos_data
            _, _, _, c1 = _merged_range_de(ws, r0, c0)
            for cc in range(c0, min(c0 + 4, ws.max_column + 1)):
                v = ws.cell(row=r0, column=cc).value
                if isinstance(v, datetime):
                    competencia_dt = v
                    break
        if competencia_dt is None:
            # modelo novo (jul/2026): a célula "DATA:" existe mas veio em
            # branco — cai pro mês/ano no nome do arquivo
            # ("..._JULHO2026_...xlsx"), igual já fazíamos pra REGIÃO:.
            m = re.search(
                r'(JANEIRO|FEVEREIRO|MAR[CÇ]O|ABRIL|MAIO|JUNHO|JULHO|AGOSTO|SETEMBRO|OUTUBRO|NOVEMBRO|DEZEMBRO)\s*(\d{4})',
                os.path.basename(caminho), re.IGNORECASE,
            )
            if m:
                mes_num = NUM_DO_MES_PT.get(_norm(m.group(1)))
                if mes_num:
                    competencia_dt = datetime(int(m.group(2)), mes_num, 1)
                    qa(f'aba "{nome_aba}": sem célula DATA: preenchida na ficha — usei {m.group(1).upper()}/{m.group(2)} do nome do arquivo')

        regiao_num = None
        if pos_regiao:
            r0, c0 = pos_regiao
            for cc in range(c0, min(c0 + 5, ws.max_column + 1)):
                v = ws.cell(row=r0, column=cc).value
                if isinstance(v, (int, float)):
                    regiao_num = int(v)
                    break
        if regiao_num is None:
            # alguns modelos de ficha (ex.: LOTE 01) não têm célula "REGIÃO:" —
            # cai pro número no nome do arquivo ("...R.01 - JULHO.xlsx" no
            # modelo antigo, "..._R13_PAVIMENTADAS.xlsx" no novo). Usa
            # negative lookahead em vez de \b: '_' é caractere de palavra,
            # então \b não marca fronteira entre "13" e "_" no nome novo.
            m = re.search(r'R\.?\s*(\d{1,2})(?!\d)', os.path.basename(caminho), re.IGNORECASE)
            if m:
                regiao_num = int(m.group(1))
                qa(f'aba "{nome_aba}": sem célula REGIÃO: na ficha — usei R{regiao_num} do nome do arquivo')

        rodovia_nome = None
        if pos_ficha:
            r0, c0 = pos_ficha
            for cc in range(c0 + 1, min(c0 + 6, ws.max_column + 1)):
                v = ws.cell(row=r0, column=cc).value
                if v:
                    rodovia_nome = str(v).strip()
                    break
        if rodovia_nome is None:
            # modelo novo (jul/2026): o título "FICHA DE INSPEÇÃO" e o nome
            # da rodovia ficam na MESMA célula mesclada (sem coluna vizinha
            # com o nome) — o nome real está na linha "TRECHO N – RODOVIA
            # ...", que sempre existe nos dois modelos.
            for r in range(1, 6):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=r, column=c).value
                    if not v:
                        continue
                    m = re.match(r'\s*TRECHO\s+\d+\s*[-–—]\s*(.+)', str(v), re.IGNORECASE)
                    if m:
                        rodovia_nome = m.group(1).strip()
                        break
                if rodovia_nome:
                    break

        if competencia_dt is None or regiao_num is None:
            print(f'  [aviso] aba "{nome_aba}": não achei DATA e/ou REGIÃO no cabeçalho — pulando')
            continue

        regiao = f'R{regiao_num}'
        competencia = f'{competencia_dt.year:04d}-{competencia_dt.month:02d}'

        # --- qual modelo de ficha é essa aba? (pavimentada x não pavimentada) ---
        r_sre, c_sre = pos_sre
        data_start_row_provisorio = _merged_range_de(ws, r_sre, c_sre)[2] + 1
        tipo_via = None
        for chave_tpl, tpl in TEMPLATES.items():
            if _achar_cabecalho(ws, tpl['chave_deteccao'], linhas=range(1, data_start_row_provisorio)):
                tipo_via = chave_tpl
                break
        if tipo_via is None:
            print(f'  [aviso] aba "{nome_aba}": não reconheci o modelo da ficha (nem pavimentada nem não pavimentada) — pulando')
            continue

        # --- localizar as colunas de cada grupo de condição do modelo detectado ---
        _, c_sentido = pos_sentido
        _, c_inicio = pos_inicio
        _, c_fim = pos_fim
        data_start_row = data_start_row_provisorio  # linha após o cabeçalho de 2 linhas

        grupos_cols = []
        for g in TEMPLATES[tipo_via]['grupos']:
            pos = _achar_cabecalho(ws, g['chave_topo'], linhas=range(1, data_start_row))
            if not pos:
                print(f'  [aviso] aba "{nome_aba}": grupo "{g["nome"]}" não encontrado')
                continue
            r0, c0 = pos
            r1, cc1, r2, cc2 = _merged_range_de(ws, r0, c0)
            cols = list(range(cc1, cc2 + 1))
            labels = [ws.cell(row=r2 + 1, column=cc).value for cc in cols]
            labels = [(re.sub(r'\s+', ' ', str(l)).strip() if l else None) for l in labels]
            grupos_cols.append({**g, 'cols': cols, 'labels': labels})

        # --- ler as linhas de dados ---
        segmentos = []
        sre_atual = None
        sentido_atual = None
        for r in range(data_start_row, ws.max_row + 1):
            v_sre = ws.cell(row=r, column=c_sre).value
            v_sentido = ws.cell(row=r, column=c_sentido).value
            v_ini = ws.cell(row=r, column=c_inicio).value
            v_fim = ws.cell(row=r, column=c_fim).value
            if v_sre not in (None, ''):
                sre_atual = str(v_sre).strip()
            if v_sentido not in (None, ''):
                sentido_atual = str(v_sentido).strip()
            if not isinstance(v_ini, (int, float)) or not isinstance(v_fim, (int, float)):
                continue
            if sre_atual is None:
                continue

            seg = {
                'sre': sre_atual,
                'sentido': sentido_atual,
                'km_ini': round(float(v_ini), 3),
                'km_fim': round(float(v_fim), 3),
                'tipo_via': tipo_via,
            }
            for g in grupos_cols:
                marcados = []
                pior_idx = None
                for i, cc in enumerate(g['cols']):
                    val = ws.cell(row=r, column=cc).value
                    if val not in (None, ''):
                        marcados.append(g['labels'][i] or f'col{i}')
                        pior_idx = i if pior_idx is None else max(pior_idx, i)
                seg[g['id']] = {
                    'marcados': marcados,
                    'severidade': pior_idx,
                    'status': (g['labels'][pior_idx] if pior_idx is not None and g['labels'][pior_idx] else None),
                }
            segmentos.append(seg)

        if not segmentos:
            continue

        segmentos = corrigir_sre_travado(segmentos, info_sre, trecho_ordem, regiao, nome_aba)

        # 'trecho' agrupa vários S.R.E. sob o mesmo lote/rodovia — equivale à
        # coluna Id da tabela de atributos do shapefile (R<n>_TRECHOS.shp) e
        # bate com o número da aba da ficha ("202", "206"...). Prioriza o que
        # vem da aba "Trechos" (autoritativo); cai pro nome da aba se faltar.
        def _trecho_de(sre):
            info = info_sre.get(_norm(sre).replace(' ', ''))
            if info and info.get('trecho_num') is not None:
                return info['trecho_num'], info.get('trecho_nome') or rodovia_nome
            # modelo novo (jul/2026) não tem a aba "Trechos" — info_sre fica
            # vazio e cai aqui pro nome da aba, que o openpyxl sempre lê como
            # STRING (mesmo quando é só "225"). Sem isso o trecho_num sai
            # inconsistente (string aqui, número quando vem de info_sre) e
            # quebra comparação estrita (===) no index.html — ex.: a aba
            # "Ordens de Serviço" parava de achar a geometria do trecho.
            try:
                return int(nome_aba), rodovia_nome
            except (TypeError, ValueError):
                return nome_aba, rodovia_nome

        segs_com_trecho = []
        for s in segmentos:
            trecho_num, trecho_nome = _trecho_de(s['sre'])
            segs_com_trecho.append({**s, 'rodovia': rodovia_nome, 'aba': nome_aba,
                                     'trecho_num': trecho_num, 'trecho_nome': trecho_nome})
        resultados_por_regiao.setdefault(regiao, {}).setdefault(competencia, []).extend(segs_com_trecho)

    saida = []
    for regiao, por_comp in resultados_por_regiao.items():
        for competencia, segs in por_comp.items():
            saida.append({'regiao': regiao, 'competencia': competencia, 'segmentos': segs})
    return saida


BASE_RODS_2023 = os.path.join(CAMADAS_DIR, 'Base_Rods_2023.shp')
_cache_base_rods_2023 = None

# Limite municipal do TO inteiro (copiado de MAPAS OSP/SHAPEFILES UTEIS —
# fonte AGM/2022) — só contexto visual, não tem relação com as regiões de
# inspeção.
LIMITES_MUNICIPAIS_SHP = os.path.join(CAMADAS_DIR, 'LimiteMunicipal_AGM_TO_2022_A.shp')


def _carregar_base_rods_2023():
    """Camada estadual complementar (todo o TO, feita pelo pessoal de campo em
    2023 — ainda não é a versão oficial final). Usada só como FALLBACK: só
    entra pros S.R.E. que faltarem no shapefile principal de cada região, sem
    substituir o que já funciona. Carregada uma vez só (mesmo arquivo pras
    6 regiões) e cacheada. Quando a versão oficial estiver pronta, a usuária
    vai avisar pra trocar — ver CLAUDE.md."""
    global _cache_base_rods_2023
    if _cache_base_rods_2023 is None:
        if os.path.exists(BASE_RODS_2023):
            _cache_base_rods_2023 = _carregar_linhas_de_shapefile(BASE_RODS_2023, 'Base_Rods_2023 (fallback)')
        else:
            _cache_base_rods_2023 = {}
    return _cache_base_rods_2023


def carregar_linhas_regiao(regiao):
    """Carrega camadas/<regiao>_TRECHOS.shp; devolve dict SRE(normalizado) ->
    dict com geometria. Só os S.R.E. que a ficha realmente pedir e não
    acharem aqui vão cair pro fallback Base_Rods_2023 (ver `linha_do_sre`)."""
    caminho = os.path.join(CAMADAS_DIR, f'{regiao}_TRECHOS.shp')
    if not os.path.exists(caminho):
        qa(f'{regiao}: shapefile {os.path.basename(caminho)} não encontrado em camadas/ — usando só o fallback Base_Rods_2023')
        return {}
    return _carregar_linhas_de_shapefile(caminho, regiao)


_avisados_fallback = set()


def linha_do_sre(linhas, sre_key, regiao):
    """Busca sre_key no dict da região; se não achar, tenta no fallback
    estadual Base_Rods_2023 (versão do pessoal de campo, ainda não oficial)
    — só avisa quando esse fallback realmente é usado, não o arquivo todo
    (e só uma vez por região+S.R.E., não por segmento de km)."""
    info = linhas.get(sre_key)
    if info is not None:
        return info, False
    fallback = _carregar_base_rods_2023()
    info = fallback.get(sre_key)
    if info is not None:
        if (regiao, sre_key) not in _avisados_fallback:
            _avisados_fallback.add((regiao, sre_key))
            qa(f'{regiao}: S.R.E. {sre_key!r} não estava no shapefile da região — usado o fallback Base_Rods_2023 '
               f'(versão do pessoal de campo, ainda não oficial)')
        return info, True
    return None, False


def _carregar_linhas_de_shapefile(caminho, label):
    """Lê UM shapefile de trechos (independente do esquema de colunas) e
    devolve dict SRE(normalizado) -> {geom_m, comprimento_m, ext_real_km}."""
    gdf_m = gpd.read_file(caminho)
    if gdf_m.crs is None:
        gdf_m = gdf_m.set_crs(EPSG_METRICO)
    else:
        gdf_m = gdf_m.to_crs(EPSG_METRICO)
    gdf_deg = gdf_m.to_crs(4326)

    col_sre = next((c for c in gdf_m.columns if _norm(c) in ('SRE', 'CODIGO')), None)
    if col_sre is None:
        qa(f'{label}: shapefile sem coluna SRE/CODIGO reconhecível (colunas: {list(gdf_m.columns)})')
        return {}

    # Os shapefiles R<n>_TRECHOS.shp NÃO seguem um esquema único — cada região
    # foi digitalizada num lote/época diferente. R11/R12/R13 usam
    # X_LONG/Y_LAT/X_FIM_LONG/Y_FIM_LAT/EXT_REAL; R1 usa START_X/START_Y/
    # END_X/END_Y/EXTENSÃO; R2 usa X_INICIO/Y_INICIO/X_FINAL/Y_FINAL; R3 usa
    # X_INICIO/Y_INICIO/X_FIM/Y_FIM. Resolve por nome normalizado (sem
    # acento/maiúscula) em vez de depender de um esquema fixo.
    def _achar_coluna(*candidatos):
        candidatos_norm = [_norm(c) for c in candidatos]
        for c in gdf_m.columns:
            if _norm(c) in candidatos_norm:
                return c
        return None

    col_x_ini = _achar_coluna('X_LONG', 'START_X', 'X_INICIO')
    col_y_ini = _achar_coluna('Y_LAT', 'START_Y', 'Y_INICIO')
    col_x_fim = _achar_coluna('X_FIM_LONG', 'END_X', 'X_FINAL', 'X_FIM')
    col_y_fim = _achar_coluna('Y_FIM_LAT', 'END_Y', 'Y_FINAL', 'Y_FIM')
    col_ext = _achar_coluna('EXT_REAL', 'EXTENSAO', 'EXTENSCAO', 'EXT_KM')
    if col_x_ini is None or col_y_ini is None or col_x_fim is None or col_y_fim is None:
        qa(f'{label}: shapefile sem colunas de coordenada início/fim reconhecíveis '
           f'(colunas: {list(gdf_m.columns)}) — não vai dar pra checar o sentido das linhas')

    linhas = {}
    for i in range(len(gdf_m)):
        sre_raw = gdf_m.iloc[i][col_sre]
        if sre_raw is None or (isinstance(sre_raw, float) and math.isnan(sre_raw)):
            continue
        sre_key = _norm(sre_raw).replace(' ', '')
        if not sre_key:
            continue
        geom_m = gdf_m.iloc[i].geometry
        if geom_m is None or geom_m.is_empty:
            continue
        if geom_m.geom_type == 'MultiLineString':
            # As partes não se tocam exatamente (emenda de digitalização, ou
            # até ramos/duplicatas de digitalização). Concatenar "na ordem
            # do arquivo" (jeito antigo) dava zigue-zague feio quando a ordem
            # não seguia o traçado real — em vez disso, encadeia sempre pela
            # PARTE MAIS PRÓXIMA de uma das pontas da linha já montada
            # (guloso: começa pela parte mais longa, vai grudando a mais
            # perto de qualquer ponta, invertendo se precisar), preservando
            # o comprimento total. Se o vão de alguma emenda for grande,
            # avisamos — pode faltar pedaço do traçado no shapefile.
            partes_todas = list(geom_m.geoms)
            n_partes_original = len(partes_todas)
            # Parte minúscula (<=50m) é vértice solto/duplicado — ruído de
            # digitalização, não estrada de verdade. Descarta ANTES de
            # encadear: se entrar na linha, o algoritmo guloso desenha um
            # "espeto" reto até ela (às vezes vários km de distância) e volta
            # — cria um laço sem sentido no mapa, mesmo o comprimento
            # descartado sendo desprezível. Mantém pelo menos a parte mais
            # longa mesmo que só sobre lixo (SRE praticamente todo ruído).
            partes_restantes = sorted(partes_todas, key=lambda p: -p.length)
            maior = partes_restantes[0]
            partes_significativas = [maior] + [p for p in partes_restantes[1:] if p.length > 50]
            n_descartadas = n_partes_original - len(partes_significativas)
            partes_restantes = list(partes_significativas)
            partes_restantes.sort(key=lambda p: -p.length)
            cadeia = list(partes_restantes.pop(0).coords)
            maior_vao = 0.0
            while partes_restantes:
                ponta_ini, ponta_fim = cadeia[0], cadeia[-1]
                melhor = None  # (distancia, indice, 'ini'|'fim', coords_da_parte_no_sentido_certo)
                for idx, p in enumerate(partes_restantes):
                    pc = list(p.coords)
                    candidatos = [
                        (( (ponta_fim[0]-pc[0][0])**2 + (ponta_fim[1]-pc[0][1])**2 )**0.5, 'fim', pc),
                        (( (ponta_fim[0]-pc[-1][0])**2 + (ponta_fim[1]-pc[-1][1])**2 )**0.5, 'fim', pc[::-1]),
                        (( (ponta_ini[0]-pc[0][0])**2 + (ponta_ini[1]-pc[0][1])**2 )**0.5, 'ini', pc[::-1]),
                        (( (ponta_ini[0]-pc[-1][0])**2 + (ponta_ini[1]-pc[-1][1])**2 )**0.5, 'ini', pc),
                    ]
                    for dist, lado, coords_certas in candidatos:
                        if melhor is None or dist < melhor[0]:
                            melhor = (dist, idx, lado, coords_certas)
                dist, idx, lado, coords_certas = melhor
                maior_vao = max(maior_vao, dist)
                if lado == 'fim':
                    cadeia.extend(coords_certas)
                else:
                    cadeia = coords_certas + cadeia
                partes_restantes.pop(idx)
            geom_m = LineString(cadeia)
            descartadas_msg = f' ({n_descartadas} minúscula(s) <=50m descartada(s), provável ruído de digitalização)' if n_descartadas else ''
            qa(f'{label}/{sre_raw}: geometria em {n_partes_original} partes desconexas no shapefile{descartadas_msg} — '
               f'encadeadas pela parte mais próxima em cada ponta (maior vão entre partes: {round(maior_vao)} m'
               f'{" — CONFERIR o shapefile, pode estar faltando um pedaço do traçado" if maior_vao > 200 else ""})')
        if sre_key in linhas:
            qa(f'{label}: SRE {sre_raw!r} duplicado no shapefile — usando a primeira ocorrência')
            continue

        # Descobre se o vértice inicial da linha corresponde ao km 0 (INÍCIO)
        # ou ao km final (FIM) do trecho, comparando com os atributos
        # X_LONG/Y_LAT (início) e X_FIM_LONG/Y_FIM_LAT (fim) — ambos em graus.
        # Se estiver invertida, vira a linha antes de cortar, senão
        # km_ini/km_fim ficam trocados. Reprojeta só o 1º vértice (já
        # fundido/reduzido a LineString única) pra graus, na hora.
        if col_x_ini and col_y_ini and col_x_fim and col_y_fim:
            ini_lonlat = (gdf_m.iloc[i].get(col_x_ini), gdf_m.iloc[i].get(col_y_ini))
            fim_lonlat = (gdf_m.iloc[i].get(col_x_fim), gdf_m.iloc[i].get(col_y_fim))
            geom_deg = gpd.GeoSeries([geom_m], crs=EPSG_METRICO).to_crs(4326).iloc[0]
            p0 = geom_deg.coords[0]
            d0 = _dist2(p0, ini_lonlat)
            d1 = _dist2(p0, fim_lonlat)
            if d0 is not None and d1 is not None and d1 < d0:
                geom_m = LineString(list(geom_m.coords)[::-1])
            elif d0 is None:
                qa(f'{label}/{sre_raw}: sem X_LONG/Y_LAT preenchido nessa linha pra checar o sentido — assumindo a ordem original')
        # se o shapefile inteiro não tem essas colunas (ex.: Base_Rods_2023),
        # já avisamos uma vez só lá em cima — não repete por feição.

        linhas[sre_key] = {
            'geom_m': geom_m,
            'comprimento_m': geom_m.length,
            'ext_real_km': (gdf_m.iloc[i].get(col_ext) if col_ext else None),
        }
    return linhas


def _dist2(a, b):
    if a[0] is None or b[0] is None or a[1] is None or b[1] is None:
        return None
    return (a[0] - b[0]) ** 2 + (a[1] - b[1]) ** 2


def cortar_segmento(linha_info, km_ini, km_fim, regiao, sre):
    geom = linha_info['geom_m']
    comprimento = linha_info['comprimento_m']

    d_ini_m = max(0.0, min(km_ini * 1000.0, comprimento))
    d_fim_m = max(0.0, min(km_fim * 1000.0, comprimento))
    if d_fim_m <= d_ini_m:
        d_fim_m = min(comprimento, d_ini_m + 0.5)

    cortado = substring(geom, d_ini_m, d_fim_m)
    if cortado.is_empty or cortado.geom_type != 'LineString' or len(cortado.coords) < 2:
        return None
    return cortado


def js_string(valor):
    return json.dumps(valor, ensure_ascii=False)


def gerar_regioes():
    """Lê camadas/R<n>_REGIÃO.shp (contorno de cada região) e escreve dados/regioes.js —
    usado só pra desenhar o limite da região selecionada no mapa (contexto visual)."""
    caminhos = sorted(glob.glob(os.path.join(CAMADAS_DIR, 'R*_REGIÃO.shp')))
    if not caminhos:
        qa('Nenhum shapefile R*_REGIÃO.shp encontrado em camadas/ — sem contorno de região no mapa')
        return
    features = []
    for caminho in caminhos:
        m = re.match(r'R(\d+)_', os.path.basename(caminho))
        if not m:
            continue
        regiao = f'R{int(m.group(1))}'
        gdf = gpd.read_file(caminho)
        gdf = gdf.to_crs(4326) if gdf.crs else gdf.set_crs(4326)
        for geom in gdf.geometry:
            if geom is None or geom.is_empty:
                continue
            features.append({'type': 'Feature', 'properties': {'regiao': regiao},
                              'geometry': mapping(geom)})
    geojson = {'type': 'FeatureCollection', 'features': features}
    with open(os.path.join(DADOS_DIR, 'regioes.js'), 'w', encoding='utf-8') as f:
        f.write('// Gerado por converter_fichas.py — não editar à mão\n')
        f.write(f'window.DADOS_REGIOES = {json.dumps(geojson, ensure_ascii=False)};\n')
    print(f'dados/regioes.js escrito com {len(features)} região(ões).')



# Tolerância maior que a da malha viária — é polígono (não precisa mostrar
# reentrância fina de fronteira), então dá pra simplificar mais sem perder
# a noção de "esse trecho tá dentro de qual município".
TOLERANCIA_SIMPLIFICACAO_MUNICIPIOS = 0.0015  # ~150 m


def gerar_limites_municipais():
    """Lê camadas/LimiteMunicipal_AGM_TO_2022_A.shp (139 municípios do TO) e
    escreve dados/limites_municipais.js — camada opcional (checkbox), só
    contorno + nome ao passar o mouse, pra ajudar a localizar em qual
    município cada trecho está."""
    if not os.path.exists(LIMITES_MUNICIPAIS_SHP):
        qa('LimiteMunicipal_AGM_TO_2022_A.shp não encontrado em camadas/ — sem limites municipais no mapa')
        return
    # encoding='utf-8' explícito: o .dbf desse shapefile vem com um .cst (não um
    # .cpg, que é o que o GDAL reconhece) indicando UTF-8 — sem isso o pyogrio
    # adivinha errado e devolve nome de município com acento em dobro
    # ("PalmeirÃ³polis" em vez de "Palmeirópolis").
    gdf = gpd.read_file(LIMITES_MUNICIPAIS_SHP, encoding='utf-8')
    gdf = gdf.to_crs(4326) if gdf.crs else gdf.set_crs(4326)
    gdf['geometry'] = gdf.geometry.simplify(TOLERANCIA_SIMPLIFICACAO_MUNICIPIOS, preserve_topology=True)

    col_nome = next((c for c in gdf.columns if _norm(c) == 'NOME'), None)
    if col_nome is None:
        qa(f'LimiteMunicipal_AGM_TO_2022_A.shp: sem coluna "nome" (colunas: {list(gdf.columns)}) — municípios sem rótulo')

    features = []
    for _, row in gdf.iterrows():
        geom = row.geometry
        if geom is None or geom.is_empty:
            continue
        props = {'nome': str(row[col_nome]).strip() if col_nome and row[col_nome] else ''}
        features.append({'type': 'Feature', 'properties': props, 'geometry': mapping(geom)})

    geojson = {'type': 'FeatureCollection', 'features': features}
    with open(os.path.join(DADOS_DIR, 'limites_municipais.js'), 'w', encoding='utf-8') as f:
        f.write('// Gerado por converter_fichas.py — não editar à mão\n')
        f.write(f'window.LIMITES_MUNICIPAIS = {json.dumps(geojson, ensure_ascii=False)};\n')
    print(f'dados/limites_municipais.js escrito com {len(features)} município(s).')


LOCALIDADES_SHP = os.path.join(CAMADAS_DIR, 'BCLocalidadePonto100_GCS.shp')


def gerar_localidades():
    """Lê camadas/BCLocalidadePonto100_GCS.shp (base cartográfica de
    localidades/povoados, mesma camada de contexto do geoportal principal —
    ver gerar_mapa.py) e escreve dados/localidades.js — pontinhos com nome
    ao passar o mouse, filtrados só pro Tocantins (coluna NM_UF)."""
    if not os.path.exists(LOCALIDADES_SHP):
        qa('BCLocalidadePonto100_GCS.shp não encontrado em camadas/ — sem localidades no mapa')
        return
    gdf = gpd.read_file(LOCALIDADES_SHP)
    gdf = gdf[gdf['NM_UF'] == 'TO'].copy()
    gdf = gdf.to_crs(4326) if gdf.crs else gdf.set_crs(4326)

    col_nome = next((c for c in ('NM_IDENTIF', 'NM_LOC_ASS', 'NOME', 'nome') if c in gdf.columns), None)
    if col_nome is None:
        qa(f'BCLocalidadePonto100_GCS.shp: sem coluna de nome reconhecível (colunas: {list(gdf.columns)})')

    features = []
    for _, row in gdf.iterrows():
        geom = row.geometry
        if geom is None or geom.is_empty:
            continue
        nome = row[col_nome] if col_nome else None
        if not nome:
            continue
        features.append({'type': 'Feature', 'properties': {'nome': str(nome).strip()}, 'geometry': mapping(geom)})

    geojson = {'type': 'FeatureCollection', 'features': features}
    with open(os.path.join(DADOS_DIR, 'localidades.js'), 'w', encoding='utf-8') as f:
        f.write('// Gerado por converter_fichas.py — não editar à mão\n')
        f.write(f'window.LOCALIDADES = {json.dumps(geojson, ensure_ascii=False)};\n')
    print(f'dados/localidades.js escrito com {len(features)} localidade(s).')


REGIOES_PONTOS_CRITICOS = ['R1', 'R2', 'R3', 'R11', 'R12', 'R13']


def _classificar_situacao(texto):
    """Reduz o texto livre da coluna de status mensal a um punhado de classes
    (pra colorir o marcador) — mesma ideia do I.C.M., só que pra pontos
    críticos."""
    t = _norm(texto or '')
    if not t or t in ('-', 'NA', 'N/A'):
        return 'Sem atualização'
    if 'RESOLV' in t or 'CONCLU' in t or 'FINALIZ' in t or 'NORMALIZ' in t or 'RECUPER' in t:
        return 'Resolvido'
    if 'EXECU' in t or 'ANDAMENTO' in t or 'OBRA' in t:
        return 'Em execução'
    if 'CRITIC' in t:
        return 'Crítico'
    return 'Outro'


MESES_ABREV_PC = {
    'JAN': 'Janeiro', 'FEV': 'Fevereiro', 'MAR': 'Março', 'ABR': 'Abril',
    'MAI': 'Maio', 'JUN': 'Junho', 'JUL': 'Julho', 'AGO': 'Agosto',
    'SET': 'Setembro', 'OUT': 'Outubro', 'NOV': 'Novembro', 'DEZ': 'Dezembro',
}


def _mes_abrev_de(texto):
    """Acha a abreviação de mês (3 letras) em qualquer posição do texto
    normalizado — usado como chave pra casar a coluna de status ('Novembro',
    'Novembro/2025'...) com a coluna do link do mapa ('Mapa Nov', 'Mapas de
    Out', 'MapaAbril' sem espaço, 'Mapa Novembro/2025'...), mesmo quando os
    dois cabeçalhos têm formato bem diferente (ano no meio, sem espaço etc.)."""
    h = _norm(texto)
    for abrev in MESES_ABREV_PC:
        if abrev in h:
            return abrev
    return None


def _mes_da_coluna_mapa(header):
    """Só conta como coluna 'Mapa de <mês>' se o cabeçalho começar com
    'MAPA' — sem isso qualquer coluna de status cujo nome contenha um mês
    (a maioria) seria confundida com coluna de link."""
    h = _norm(header)
    if not h.startswith('MAPA'):
        return None
    return _mes_abrev_de(header)


def _ler_planilha_pontos_criticos():
    """Lê Controle Pontos Críticos .xlsx (uma aba por região) e devolve dict
    regiao -> {PONTO(int): {...}}. O layout varia um pouco de região pra
    região (meses diferentes, nome da coluna de situação final diferente) —
    resolve pelo nome normalizado da coluna, igual ao resto do conversor."""
    if not os.path.exists(PLANILHA_PONTOS_CRITICOS):
        qa(f'Pontos Críticos: planilha não encontrada em {PLANILHA_PONTOS_CRITICOS} '
           f'— camada vai ficar sem status/descrição, só a geometria')
        return {}
    wb = openpyxl.load_workbook(PLANILHA_PONTOS_CRITICOS, data_only=True)
    saida = {}
    for nome_aba in wb.sheetnames:
        m = re.match(r'REGI[ÃA]O\s*(\d+)', nome_aba, re.IGNORECASE)
        if not m:
            continue
        regiao = f'R{int(m.group(1))}'
        ws = wb[nome_aba]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        header_norm = [_norm(h) if h else '' for h in header]

        def _achar(*candidatos):
            candidatos_norm = [_norm(c) for c in candidatos]
            for i, h in enumerate(header_norm):
                if h in candidatos_norm:
                    return i + 1  # coluna 1-based
            return None

        col_ponto = _achar('PONTO')
        col_rodovia = _achar('RODOVIA')
        col_trecho = _achar('TRECHO CORRESPONDENTE', 'TRECHO')
        col_situacao = _achar('STATUS FINAL / SITUACAO', 'SITUACAO ATUAL', 'STATUS FINAL/SITUACAO')
        if col_ponto is None:
            qa(f'Pontos Críticos {regiao}: aba {nome_aba!r} sem coluna PONTO reconhecível — pulada')
            continue

        # Colunas de mês: tudo entre "Trecho Correspondente" e a coluna de
        # situação final são status pontuais (Crítico/Em execução/-), na
        # ordem em que a usuária foi preenchendo — dá a evolução do ponto.
        col_meses = []
        if col_trecho and col_situacao and col_situacao > col_trecho:
            for c in range(col_trecho + 1, col_situacao):
                nome_mes = header[c - 1]
                if nome_mes:
                    col_meses.append((c, str(nome_mes).strip(), _mes_abrev_de(nome_mes)))

        # Colunas "Mapa de <mês>" (depois da situação final) — o texto do
        # link é sempre "Ver no Mapa", o que interessa é o hyperlink em si:
        # abre a ficha do mês em PDF (mapa de localização + fotos, hospedada
        # no Drive da usuária, privada — por isso só link, nunca embutimos
        # a imagem direto).
        col_mapas_por_mes = {}
        if col_situacao:
            for c in range(col_situacao + 1, ws.max_column + 1):
                abrev = _mes_da_coluna_mapa(header[c - 1])
                if abrev:
                    col_mapas_por_mes[abrev] = c

        pontos = {}
        for r in range(2, ws.max_row + 1):
            id_ponto = ws.cell(row=r, column=col_ponto).value
            if id_ponto is None:
                continue
            try:
                id_ponto = int(id_ponto)
            except (TypeError, ValueError):
                continue
            historico = []
            for c, nome_mes, abrev_mes in col_meses:
                v = ws.cell(row=r, column=c).value
                v = str(v).strip() if v is not None else ''
                mapa_url = None
                col_mapa = col_mapas_por_mes.get(abrev_mes) if abrev_mes else None
                if col_mapa:
                    cell_mapa = ws.cell(row=r, column=col_mapa)
                    if cell_mapa.hyperlink:
                        mapa_url = cell_mapa.hyperlink.target
                if (v and v != '-') or mapa_url:
                    historico.append({'mes': nome_mes, 'status': v or '-', 'mapa_url': mapa_url})
            ultimo_status = next((h['status'] for h in reversed(historico) if h['status'] and h['status'] != '-'), '-')
            situacao_bruta = ultimo_status
            descricao = ws.cell(row=r, column=col_situacao).value if col_situacao else None
            descricao = str(descricao).strip() if descricao else ''
            rodovia = ws.cell(row=r, column=col_rodovia).value if col_rodovia else None
            trecho = ws.cell(row=r, column=col_trecho).value if col_trecho else None
            pontos[id_ponto] = {
                'rodovia': str(rodovia).strip() if rodovia else '',
                'trecho': str(trecho).strip() if trecho else '',
                'situacao_classe': _classificar_situacao(situacao_bruta),
                'situacao_bruta': situacao_bruta,
                'descricao': descricao,
                'historico': historico,
            }
        saida[regiao] = pontos
    return saida


def gerar_pontos_criticos():
    """Lê camadas/R<n>_Pontos_Criticos.shp (geometria) + Controle Pontos
    Críticos .xlsx (status/descrição — planilha do outro projeto) e escreve
    dados/pontos_criticos.js — camada opcional (checkbox no mapa) com os
    pontos de erosão/bueiro/talude etc. já levantados em campo."""
    dados_planilha = _ler_planilha_pontos_criticos()
    features = []
    for regiao in REGIOES_PONTOS_CRITICOS:
        caminho = os.path.join(CAMADAS_DIR, f'{regiao}_Pontos_Criticos.shp')
        if not os.path.exists(caminho):
            continue
        gdf = gpd.read_file(caminho)
        gdf = gdf.to_crs(4326) if gdf.crs else gdf.set_crs(4326)
        col_id = next((c for c in gdf.columns if _norm(c) == 'ID'), None)
        if col_id is None:
            qa(f'Pontos Críticos {regiao}: shapefile sem coluna Id — pulado')
            continue
        pontos_regiao = dados_planilha.get(regiao, {})
        for _, row in gdf.iterrows():
            geom = row.geometry
            if geom is None or geom.is_empty:
                continue
            ponto_id = row[col_id]
            try:
                ponto_id = int(ponto_id)
            except (TypeError, ValueError):
                continue
            info = pontos_regiao.get(ponto_id)
            if info is None:
                qa(f'Pontos Críticos {regiao}: ponto {ponto_id} tem geometria mas não está na planilha '
                   f'— sem status/descrição')
                info = {'rodovia': '', 'trecho': '', 'situacao_classe': 'Sem atualização',
                        'situacao_bruta': '-', 'descricao': '', 'historico': []}
            # Ponto já recuperado/resolvido — não entra no mapa (a usuária só
            # quer ver o que ainda precisa de atenção).
            if info['situacao_classe'] == 'Resolvido':
                continue
            props = {'regiao': regiao, 'id': ponto_id}
            props.update(info)
            features.append({'type': 'Feature', 'properties': props, 'geometry': mapping(geom)})

    geojson = {'type': 'FeatureCollection', 'features': features}
    with open(os.path.join(DADOS_DIR, 'pontos_criticos.js'), 'w', encoding='utf-8') as f:
        f.write('// Gerado por converter_fichas.py — não editar à mão\n')
        f.write(f'window.PONTOS_CRITICOS = {json.dumps(geojson, ensure_ascii=False)};\n')
    print(f'dados/pontos_criticos.js escrito com {len(features)} ponto(s) crítico(s).')


def main():
    os.makedirs(DADOS_DIR, exist_ok=True)
    gerar_regioes()
    gerar_limites_municipais()
    gerar_localidades()
    gerar_pontos_criticos()
    # recursivo: a usuária organiza em subpastas tipo fichas/LOTE 04/JULHO/...
    arquivos = sorted(glob.glob(os.path.join(FICHAS_DIR, '**', '*.xlsx'), recursive=True))
    arquivos = [a for a in arquivos if not os.path.basename(a).startswith('~$')]
    if not arquivos:
        print(f'Nenhuma ficha .xlsx encontrada em {FICHAS_DIR}')
        return

    por_regiao_competencia = OrderedDict()  # (regiao,competencia) -> lista de segmentos crus

    for caminho in arquivos:
        try:
            saida = parse_ficha(caminho)
        except Exception as e:
            qa(f'{os.path.basename(caminho)}: falha ao ler ({e}) — arquivo ignorado')
            continue
        for bloco in saida:
            chave = (bloco['regiao'], bloco['competencia'])
            por_regiao_competencia.setdefault(chave, []).extend(bloco['segmentos'])

    cache_linhas = {}
    manifest = []

    for (regiao, competencia), segmentos in por_regiao_competencia.items():
        print(f'\n=== {regiao} — competência {competencia} — {len(segmentos)} segmento(s) na ficha ===')
        if regiao not in cache_linhas:
            cache_linhas[regiao] = carregar_linhas_regiao(regiao)
        linhas = cache_linhas[regiao]

        features = []
        sre_nao_achados = set()
        ultimo_fim_por_sre = {}

        for seg in segmentos:
            sre_key = _norm(seg['sre']).replace(' ', '')
            linha_info, veio_do_fallback = linha_do_sre(linhas, sre_key, regiao)
            if linha_info is None:
                sre_nao_achados.add(seg['sre'])
                continue
            geom_cortada = cortar_segmento(linha_info, seg['km_ini'], seg['km_fim'], regiao, seg['sre'])
            if geom_cortada is None:
                qa(f'{regiao}/{seg["sre"]}: km {seg["km_ini"]}-{seg["km_fim"]} virou geometria vazia — pulado')
                continue
            ultimo_fim_por_sre[sre_key] = max(ultimo_fim_por_sre.get(sre_key, 0), seg['km_fim'])

            props = {
                'regiao': regiao,
                'competencia': competencia,
                'competencia_label': f"{MESES_PT[int(competencia.split('-')[1])]}/{competencia.split('-')[0]}",
                'sre': seg['sre'],
                'sentido': seg['sentido'],
                'tipo_via': seg.get('tipo_via'),
                'rodovia': seg.get('rodovia'),
                'trecho_num': seg.get('trecho_num'),
                'trecho_nome': seg.get('trecho_nome'),
                'km_ini': seg['km_ini'],
                'km_fim': seg['km_fim'],
                # geometria veio do fallback estadual (Base_Rods_2023, ainda não
                # oficial) em vez do shapefile da própria região — o mapa marca
                # esses segmentos com um traço diferente.
                'fallback': veio_do_fallback,
            }
            # só grava o grupo se ele existir na ficha dessa via (pavimentada
            # tem 5 grupos, não pavimentada tem 2 — o mapa filtra cada camada
            # pelas features que realmente têm aquela chave em properties).
            for g in GRUPOS_TODOS:
                info_g = seg.get(g['id'])
                if info_g is None:
                    continue
                props[g['id']] = {
                    'status': info_g.get('status'),
                    'severidade': info_g.get('severidade'),
                    'marcados': info_g.get('marcados', []),
                }
            props['icm'] = calcular_icm(props)

            features.append({
                'type': 'Feature',
                'geometry': {'type': 'LineString', 'coords_m': list(geom_cortada.coords)},
                'properties': props,
            })

        if sre_nao_achados:
            qa(f'{regiao}/{competencia}: {len(sre_nao_achados)} S.R.E. da ficha não encontrado(s) no shapefile: {sorted(sre_nao_achados)}')

        for sre_key, fim in ultimo_fim_por_sre.items():
            ext_real = linhas.get(sre_key, {}).get('ext_real_km')
            if ext_real and abs(fim - ext_real) > max(0.3, ext_real * 0.03):
                qa(f'{regiao}/{competencia}: SRE {sre_key} — km final inspecionado ({fim}) difere da extensão do shapefile ({round(ext_real,3)})')

        if not features:
            print(f'  nenhum segmento georreferenciado para {regiao}/{competencia} — nada gerado')
            continue

        # reprojeta em lote (métrico -> WGS84) pra ficar rápido
        transformer = None
        try:
            from pyproj import Transformer
            transformer = Transformer.from_crs(EPSG_METRICO, 4326, always_xy=True)
        except Exception as e:
            qa(f'pyproj indisponível ({e}) — abortando')
            return

        for feat in features:
            coords_m = feat['geometry'].pop('coords_m')
            xs = [p[0] for p in coords_m]
            ys = [p[1] for p in coords_m]
            lons, lats = transformer.transform(xs, ys)
            feat['geometry']['coordinates'] = [[round(lo, 6), round(la, 6)] for lo, la in zip(lons, lats)]

        geojson = {'type': 'FeatureCollection', 'features': features}
        nome_arquivo = f'insp_{regiao}_{competencia}.js'
        var_nome = f"DADOS_INSPECAO_{regiao}_{competencia.replace('-', '_')}"
        conteudo = (
            f"// Gerado por converter_fichas.py — não editar à mão\n"
            f"window.DADOS_INSPECAO = window.DADOS_INSPECAO || {{}};\n"
            f"window.DADOS_INSPECAO[{js_string(regiao)}] = window.DADOS_INSPECAO[{js_string(regiao)}] || {{}};\n"
            f"window.DADOS_INSPECAO[{js_string(regiao)}][{js_string(competencia)}] = {json.dumps(geojson, ensure_ascii=False)};\n"
        )
        with open(os.path.join(DADOS_DIR, nome_arquivo), 'w', encoding='utf-8') as f:
            f.write(conteudo)
        print(f'  -> dados/{nome_arquivo} ({len(features)} segmentos)')

        ext_total = sum(
            (f['properties']['km_fim'] - f['properties']['km_ini']) for f in features
        )
        manifest.append({
            'regiao': regiao,
            'competencia': competencia,
            'competencia_label': features[0]['properties']['competencia_label'],
            'arquivo': nome_arquivo,
            'n_segmentos': len(features),
            'ext_km': round(ext_total, 1),
        })

    manifest.sort(key=lambda m: (m['regiao'], m['competencia']))
    grupos_saida = [{'id': g['id'], 'nome': g['nome']} for g in GRUPOS_TODOS]
    with open(os.path.join(DADOS_DIR, 'manifest.js'), 'w', encoding='utf-8') as f:
        f.write('// Gerado por converter_fichas.py — não editar à mão\n')
        f.write(f'window.MANIFEST_INSPECAO = {json.dumps(manifest, ensure_ascii=False, indent=2)};\n')
        f.write(f'window.GRUPOS_INSPECAO = {json.dumps(grupos_saida, ensure_ascii=False, indent=2)};\n')
    print(f'\ndados/manifest.js escrito com {len(manifest)} conjunto(s) região/competência.')

    with open(os.path.join(BASE, 'relatorio_qualidade.txt'), 'w', encoding='utf-8') as f:
        if qa_msgs:
            f.write('\n'.join(qa_msgs) + '\n')
        else:
            f.write('Nenhum problema encontrado.\n')
    print(f'relatorio_qualidade.txt: {len(qa_msgs)} observação(ões).')

    # Mesmo conteúdo do relatorio_qualidade.txt, mas em dados/ (não é
    # gitignored, então vai junto quando publicar) — o index.html usa isso
    # pra mostrar um painel de avisos na aba "Visão Geral", em vez do
    # relatório só existir como .txt que ninguém vê depois de publicado.
    with open(os.path.join(DADOS_DIR, 'avisos.js'), 'w', encoding='utf-8') as f:
        f.write('// Gerado por converter_fichas.py — não editar à mão\n')
        f.write(f'window.AVISOS_QUALIDADE = {json.dumps(qa_msgs, ensure_ascii=False, indent=2)};\n')
    print(f'dados/avisos.js: {len(qa_msgs)} aviso(s).')


if __name__ == '__main__':
    main()
